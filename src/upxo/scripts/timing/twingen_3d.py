import time
import numpy as np
import os
from upxo.ggrowth.mcgs import mcgs
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
# #########################################################################
result_folder = r'C:\Development\M2MatMod\upxo_packaged\upxo_private\documentation'
result_file = r'\twin_inclusion_profiling_results_01.xlsx'
result_file = result_folder + result_file

if os.path.exists(result_file):
    wb = load_workbook(result_file)
else:
    wb = Workbook()
# #########################################################################
data_names = ['ng', 'tgsgen', 'ttwininc', 'avggs', 'nthg', 'tvf', 'ntwins']
# #########################################################################
sheet_prefix = 'S32_'
sheet_suffix = '_a'
domains = ['25', '50', '75', '100', '125', '150', '175', '200', '250']
# #########################################################################
ref = {'25': {'input_dashboard': 'a_01.xls', 'ntrials': 10, 'ninstances': 8},
       '50': {'input_dashboard': 'a_02.xls', 'ntrials': 8, 'ninstances': 7},
       '75': {'input_dashboard': 'a_03.xls', 'ntrials': 6, 'ninstances': 6},
       '100': {'input_dashboard': 'a_04.xls', 'ntrials': 5, 'ninstances': 5},
       '125': {'input_dashboard': 'a_05.xls', 'ntrials': 5, 'ninstances': 4},
       '150': {'input_dashboard': 'a_06.xls', 'ntrials': 5, 'ninstances': 4},
       '175': {'input_dashboard': 'a_07.xls', 'ntrials': 4, 'ninstances': 4},
       '200': {'input_dashboard': 'a_08.xls', 'ntrials': 2, 'ninstances': 4},
       '250': {'input_dashboard': 'a_09.xls', 'ntrials': 2, 'ninstances': 3}}
# #########################################################################
# #########################################################################
for domain_name in domains:
    sheet_name = sheet_prefix + domain_name + sheet_suffix
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)
    # --------------------------------
    ntrials = ref[domain_name]['ntrials']
    ninstances = ref[domain_name]['ninstances']
    # --------------------------------
    SIM_TIMES = np.array([None for nt in range(ntrials)])
    AVG_GS_NVOX = np.array([None for nt in range(ntrials)])
    TWIN_INCLUSION_TIMES_avg = np.array([None for nt in range(ntrials)])
    NGRAINS = np.array([None for nt in range(ntrials)])
    NGRAINS_HOSTS = np.array([[None for ni in range(ninstances)]
                              for nt in range(ntrials)])
    NTWINS = np.array([[None for ni in range(ninstances)]
                       for nt in range(ntrials)])
    TWIN_VOLUME_FRACTION = np.array([[None for ni in range(ninstances)]
                                     for nt in range(ntrials)])
    for nt in range(ntrials):
        print(40*'\.', '\n', 40*'\.')
        print(f'Simulation repeat number {nt+1} out of {ntrials}')
        print(40*'\.', '\n', 40*'\.')
        # ---------------------------------
        start_time = time.time()
        # ---------------------------------
        pxt = mcgs(input_dashboard=ref[domain_name]['input_dashboard'])
        pxt.simulate(verbose=False)
        # ---------------------------------
        tslice = 49
        gstslice = pxt.gs[tslice]
        gstslice.char_morphology_of_grains(label_str_order=1, find_grain_voxel_locs=True,
                                           find_spatial_bounds_of_grains=True, force_compute=True)
        # ---------------------------------
        elapsed_time_simulation = time.time() - start_time
        # #########################################################################
        gstslice.set_mprop_volnv()
        AVG_GS_NVOX[nt] = np.mean(list(gstslice.mprop['volnv'].values()))
        # ###################################`######################################
        start_time = time.time()
        mprops = {'volnv': {'use': True, 'reset': False, 'k': [.02, 1.0], 'min_vol': 4},
                  'rat_sanv_volnv': {'use': True, 'reset': False, 'k': [0.0, .8], 'sanv_N': 26}}
        twspec = {'n': [5, 10, 3], 'tv': np.array([5, -3.5, 5]),
                  'dlk': np.array([1.0, -1.0, 1.0]), 'dnw': np.array([0.5, 0.5, 0.5]),
                  'dno': np.array([0.5, 0.5, 0.5]), 'tdis': 'normal',
                  'tpar': {'loc': 1.12, 'scale': 0.25, 'val': 1},
                  'vf': [0.05, 1.00], 'sep_bzcz': False}
        twgenspec = {'seedsel': 'random_gb', 'K': 10, 'bidir_tp': False, 'checks': [True, True]}
        gstslice.instantiate_twins(ninstances=ninstances, base_gs_name_prefix='twin.',
                                   twin_setup={'nprops': 2, 'mprops': mprops},
                                   twspec=twspec, twgenspec=twgenspec)
        elapsed_time_twin_inclusion = time.time() - start_time
        # #########################################################################
        SIM_TIMES[nt] = elapsed_time_simulation
        TWIN_INCLUSION_TIMES_avg[nt] = elapsed_time_twin_inclusion/ninstances
        NGRAINS[nt] = gstslice.n
        for ni in range(ninstances):
            TWIN_VOLUME_FRACTION[nt][ni] = gstslice.fdb['twin.'+str(ni)]['data']['twin_vf_total']
            NTWINS[nt][ni] = np.array(list(gstslice.fdb['twin.'+str(ni)]['data']['twin_map_g_nt'].values())).sum()
            NGRAINS_HOSTS[nt][ni] = len(gstslice.fdb['twin.'+str(ni)]['data']['twin_map_g_nt'].keys())
    # #########################################################################
    NGRAINS = np.expand_dims(NGRAINS, axis=1)
    SIM_TIMES = np.expand_dims(SIM_TIMES, axis=1)
    TWIN_INCLUSION_TIMES_avg = np.expand_dims(TWIN_INCLUSION_TIMES_avg, axis=1)
    AVG_GS_NVOX = np.expand_dims(AVG_GS_NVOX, axis=1)

    varmap = {'ng': NGRAINS,
              'tsim': SIM_TIMES,
              'ttwinc': TWIN_INCLUSION_TIMES_avg,
              'nvox': AVG_GS_NVOX,
              'nthg': NGRAINS_HOSTS,
              'vftw': TWIN_VOLUME_FRACTION,
              'ntw': NTWINS}

    current_row = 1
    for header, array in varmap.items():
        n_rows, n_cols = array.shape
        # Write header row with a "Trial" label in the first column.
        ws.cell(row=current_row, column=1, value="Trial")
        for col in range(n_cols):
            col_header = f"{header}_{col+1}" if n_cols > 1 else header
            ws.cell(row=current_row, column=col+2, value=col_header)
        # Write data rows with trial names in the first column.
        for i in range(n_rows):
            ws.cell(row=current_row + 1 + i, column=1, value=f"trial{i+1}")
            for j in range(n_cols):
                ws.cell(row=current_row + 1 + i, column=j+2, value=float(array[i, j]))
        # Add two blank rows between each block of data.
        current_row += n_rows + 3
    # #########################################################################
wb.save(result_file)
# #########################################################################
def LOAD_FROM_CELL_RANGE(workbook, sheet_name, col_start, row_start, col_end, row_end):
    """
    Extracts data from a specified cell range within a worksheet of an Excel workbook.

    Parameters:
        workbook  : An openpyxl Workbook object.
        sheet_name: The name of the worksheet to extract data from.
        col_start : The starting column letter (e.g., 'B').
        row_start : The starting row number (e.g., 2).
        col_end   : The ending column letter (e.g., 'B').
        row_end   : The ending row number (e.g., 6).

    Returns:
        A list of lists, where each inner list contains the values from one row of the specified range.
        Returns an empty list if the sheet is not found or if an error occurs.
    """
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        print(f"Warning: Sheet '{sheet_name}' not found in the workbook.")
        return []

    try:
        '''
        Convert the column letters for the starting and ending columns into numeric indices.
        Openpyxl uses these numeric indices for column access.
        '''
        col_start_idx = column_index_from_string(col_start)
        col_end_idx = column_index_from_string(col_end)

        data = []
        '''
        Iterate over each row in the specified range using the numeric indices.
        For each row, extract the value from every cell and compile the results into a list.
        '''
        for row in sheet.iter_rows(min_row=row_start, max_row=row_end,
                                   min_col=col_start_idx, max_col=col_end_idx):
            row_data = [cell.value for cell in row]
            data.append(row_data)

        ''' Return the complete data extracted from the specified cell range. '''
        return data

    except Exception as e:
        '''
        If an error occurs during extraction (for example, due to an invalid range),
        print an error message and return an empty list.
        '''
        print(f"Error reading cell range: {e}")
        return []


def plot_data_for_dn(dn, DATA_mean, DATA_std, XC, domains, figsize):
    '''
    Plots DATA_mean values across all XC values over a range of domains for a given dn.
    Error bars are added using the corresponding DATA_std values.

    Parameters:
      dn        : A string representing the data name (e.g., 'ng', 'tsim', etc.)
      DATA_mean : A nested dictionary where data is accessed as DATA_mean[xc][dn][domain]
      DATA_std  : A nested dictionary where data is accessed as DATA_std[xc][dn][domain]
      XC        : A list of xc keys (e.g., ['s05', 's10', 's32'])
      domains   : A list of domain keys (e.g., ['25', '50', '75', '100', ...])
    '''

    # Convert the domain keys (strings) to numeric values for plotting on the x-axis.
    domain_values = [float(d) for d in domains]

    # Create a new figure with a specified size.
    plt.figure(figsize=figsize)

    # Loop through each xc value to plot its corresponding data series.
    for xc in XC:
        # Retrieve the mean values and standard deviation values for the current xc and given dn.
        mean_values = [DATA_mean[xc][dn][d] for d in domains]
        std_values  = [DATA_std[xc][dn][d] for d in domains]

        # Plot the data with error bars using the errorbar function.
        # 'marker' specifies the marker style; 'capsize' adds caps to the error bars.
        plt.errorbar(domain_values, mean_values, yerr=std_values, marker='o', capsize=5, label=xc)

    # Label the axes and add a title.
    plt.xlabel(f'Domain size')
    plt.ylabel(f'mean( {dn} )')
    # plt.title(f'Plot of DATA_mean for "{dn}" across Domains and XC values')

    # Add a legend to distinguish between different XC values.
    plt.legend(title='XC')

    # Add a grid for better readability.
    plt.grid(True)

    plt.tight_layout()

    # Display the plot.
    plt.show()
# #########################################################################
result_folder = r'C:\Development\M2MatMod\upxo_packaged\upxo_private\documentation'
result_file = r'\twin_inclusion_profiling_results_01.xlsx'
result_file = result_folder + result_file

workbook = load_workbook(result_file)

# Excel cell details: ntrials col, ninstances col, row start, row end
# This is in sheet: 'Sheet'
XC = {'s05': ['h', 'i', 13, 21],
      's10': ['v', 'w', 13, 21],
      's32': ['aj', 'ak', 13, 21],
      }
# INformation to fnd ut the data sheet names
sheet_prefixes = ['S05_', 'S10_', 'S32_']
domains = ['25', '50', '75', '100', '125', '150', '175', '200', '250']
sheet_suffix = '_a'
# Names of data columns
cols = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
        'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
start_col_name = 'b'
start_col_id = cols.index(start_col_name)
starting_row = 2
row_spacing = 2  # Number of empty rtows between two data sets

data_names = ['ng', 'tsim', 'ttwinc', 'nvox', 'nthg', 'vftw', 'ntw']
data_nums = range(len(data_names))

DATA_mean = {xc: {dn: {domain: None for domain in domains}
                  for dn in data_names} for xc in XC.keys()}
DATA_std = {xc: {dn: {domain: None for domain in domains}
                 for dn in data_names} for xc in XC.keys()}
for xc, sheet_prefix in zip(XC.keys(), sheet_prefixes):

    ntrials = np.array(LOAD_FROM_CELL_RANGE(workbook, 'Sheet',
                                            XC[xc][0],  # column start
                                            XC[xc][2],  # Row start
                                            XC[xc][0],  # Column end
                                            XC[xc][3]  # row end
                                            )).squeeze()

    ninstances = np.array(LOAD_FROM_CELL_RANGE(workbook, 'Sheet',
                                               XC[xc][1],  # column start
                                               XC[xc][2],  # Row start
                                               XC[xc][1],  # Column end
                                               XC[xc][3]  # row end
                                               )).squeeze()

    for domain_i, domain in enumerate(domains, start=0):
        sheet_name = sheet_prefix + domain + sheet_suffix
        ntr, ninst = ntrials[domain_i], ninstances[domain_i]

        for dc, dn in enumerate(data_names, start=1):

            if dc == 1:
                row_start = starting_row
                row_end = row_start + ntr*dc - 1
            else:
                row_start = row_end + row_spacing + 1 + 1
                row_end = row_start + ntr - 1

            col_start = start_col_name

            if dn in ('ng', 'tsim', 'ttwinc', 'nvox'):
                col_end = start_col_name
                data = np.array(LOAD_FROM_CELL_RANGE(workbook,
                                                     sheet_name,
                                                     col_start,
                                                     row_start,
                                                     col_end,
                                                     row_end)).squeeze()
            elif dn in ('nthg', 'vftw', 'ntw'):
                col_end = cols[start_col_id+ninst-1]
                data = np.array(LOAD_FROM_CELL_RANGE(workbook,
                                                     sheet_name,
                                                     col_start,
                                                     row_start,
                                                     col_end,
                                                     row_end))

            DATA_mean[xc][dn][domain] = data.mean()
            DATA_std[xc][dn][domain] = data.std()

            # data_range = [dn, col_start, row_start, col_end, row_end]
print(40*'-', DATA_mean)


# DATA_mean['s05']['ng']['25']

data_names = ['ng', 'tsim', 'ttwinc', 'nvox', 'nthg', 'vftw', 'ntw']

plot_data_for_dn(data_names[6], DATA_mean, DATA_std, XC, domains, figsize=(6, 3))
