# -*- coding: utf-8 -*-
"""
Created on Thu May 23 09:36:32 2024

@author: rg5749
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

"""
NOTE: DO NOT CHANGE SETTINGS, IN THE INTEREST OF TOP EXPLANATIONS.
"""
def write_array_to_excel(array, file_name, sheet_name, start_cell):
    """Writes a 2D NumPy array to an Excel file.

    Args:
        array: The 2D NumPy array to write.
        file_name: The name of the Excel file (existing or new).
        sheet_name: The name of the sheet to write to.
        start_cell: The starting cell address (e.g., "A1").
    """
    df = pd.DataFrame(array)
    try:
        wb = load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    # Get the starting row and column from the start_cell
    start_row = int(start_cell[1:]) - 1  # Convert column letter to index (0-based)
    start_col = ord(start_cell[0].upper()) - ord('A')
    # Write the DataFrame to the sheet, starting at the specified cell
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for col_idx, value in enumerate(row):
            ws.cell(row=start_row + row_idx + 1, column=start_col + col_idx + 1, value=value)
    wb.save(file_name)

def read_excel_range(file_path, sheet_name, cell_range):
    """
    Reads a specific cell range from an Excel file into a pandas DataFrame.

    Parameters:
    file_path (str): The path to the Excel file.
    sheet_name (str): The sheet name to read from.
    cell_range (str): The cell range to read (e.g., 'B2:D10').

    Returns:
    pd.DataFrame: DataFrame containing the data from the specified cell range.
    """
    # Read the entire sheet to access the specific range
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    # Convert the cell range to start and end indices
    start_cell, end_cell = cell_range.split(':')
    start_row = int(start_cell[1:]) - 1
    end_row = int(end_cell[1:])
    start_col = ord(start_cell[0].upper()) - ord('A')
    end_col = ord(end_cell[0].upper()) - ord('A') + 1
    # Extract the specified range
    data = df.iloc[start_row:end_row, start_col:end_col]
    # Set the first row as column names
    data.columns = data.iloc[0]
    data = data[1:]
    data.reset_index(drop=True, inplace=True)
    return data
